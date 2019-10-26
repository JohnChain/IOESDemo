# -*- coding: utf-8 -*-
#!/usr/local/bin/python
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class BaseDump():
    def __init__(self, destFile, mapSheetName, ):
        self.destFile = destFile
        self.wb = Workbook()
        # 已ObjectType为key的字典，value为 sheet 句柄
        self.wsDict = {}
        for key, value in mapSheetName.items():
            self.wsDict[key] = self.wb.create_sheet(title=value, index=0)

    def save(self):
        self.wb.save(filename = self.destFile)

    # 在指定sheet里插入一行
    # objType: 目标类型，将根据此类型映射到对应类型的sheet
    # listOneRow: 一行数据，（建议list，其他可迭代(dict)未尝试）
    def insert(self, objType, listOneRow):
        ws = self.wsDict[objType]
        ws.append(listOneRow)
    def insertImage(self, objType, img):
        ws = self.wsDict[objType]
        ws.add_image(img)

    def getCurrentRow(self, objType):
        return self.wsDict[objType].max_row

if __name__ == "__main__":
    TYPE_PERSON     = "1"
    TYPE_CAR        = "2"
    TYPE_FACE       = "3"
    TYPE_BIKE       = "4"

    ObjectType = {
        TYPE_FACE: "人脸",
        TYPE_BIKE: "非机动车",
        TYPE_PERSON: "行人",
        TYPE_CAR: "机动车",
    }

    dumper = BaseDump("empty_book2.xlsx", ObjectType)
    dumper.save()