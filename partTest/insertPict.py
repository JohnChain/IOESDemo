# -*- coding: utf-8 -*-
#!/usr/local/bin/python
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# excel_address = r"temp.xlsx"
# wb = Workbook()
# sht = wb.worksheets[0]

# img_address_1 = r"temp.png"
# img = Image(img_address_1)
# sht.add_image(img, 'A1')

# sht.column_dimensions['K'].width = 20.0
# sht.row_dimensions[1].height = 40.0

# img_address_2 = r"temp.png"
# img = Image(img_address_2)
# img.width = 19.0
# img.height = 39.0

# sht.add_image(img, 'K1')

# wb.save(excel_address)
import openpyxl

wb = openpyxl.Workbook()
ws = wb.worksheets[0]
img = openpyxl.drawing.image.Image('sample.jpg')
img.height = 50
img.width = 50
img.anchor = 'A1'

ws.add_image(img)
print("max_row: %d max_column:%d min_row: %d min_column: %d" %(ws.max_row, ws.max_column, ws.min_row, ws.min_column))
img = openpyxl.drawing.image.Image('sample.jpg')
img.height = 50
img.width = 50
img.anchor = 'A2'
ws.add_image(img)
print("max_row: %d max_column:%d min_row: %d min_column: %d" %(ws.max_row, ws.max_column, ws.min_row, ws.min_column))

img = openpyxl.drawing.image.Image('sample.jpg')
img.height = 50
img.width = 50
img.anchor = 'B1'
ws.add_image(img)
print("max_row: %d max_column:%d min_row: %d min_column: %d" %(ws.max_row, ws.max_column, ws.min_row, ws.min_column))
wb.save('out.xlsx')
ws = wb.worksheets[0]
print("max_row: %d max_column:%d min_row: %d min_column: %d" %(ws.max_row, ws.max_column, ws.min_row, ws.min_column))